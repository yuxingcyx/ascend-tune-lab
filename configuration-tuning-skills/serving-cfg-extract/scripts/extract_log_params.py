#!/usr/bin/env python3
"""
serving-cfg-extract: 从服务化日志启动阶段提取 non-default args 参数并生成 Excel 报告。

用法:
    python3 extract_log_params.py <日志文件路径>

输出:
    桌面的 .xlsx 文件
"""

import re
import json
import ast
import sys
from pathlib import Path

# ============================================================
# 关注的参数白名单
# ============================================================
INTERESTED_KEYS = {
    # 1. 模型路径
    "model",
    # 2. 量化方式
    "quantization",
    # 3. 并行策略
    "tensor_parallel_size", "enable_expert_parallel",
    "data_parallel_size", "pipeline_parallel_size",
    # 4. 配置参数
    "max_model_len", "max_num_batched_tokens", "max_num_seqs",
    # 5. 特性开关
    "enable_prefix_caching", "enable_chunked_prefill",
    "cudagraph_mode", "cudagraph_capture_size",
    "multistream_overlap_shared_expert",
    "enable_npugraph_ex", "recompute_scheduler_enable",
    "enable_flashcomm1", "enable_fused_mc2", "enable_dsa_cp",
    "speculative_config", "num_spec_tokens",
}

# Wildcard prefixes - any key starting with these is kept
WILDCARD_PREFIXES = ("enable_", "fuse_", "cudagraph_", "multistream_", "recompute_")
WILDCARD_PREFIX_PAT = "|".join(WILDCARD_PREFIXES)


def is_interested_key(key: str) -> bool:
    """判断一个参数名是否被关注。"""
    if key in INTERESTED_KEYS:
        return True
    last_part = key.split('.')[-1]
    if last_part in INTERESTED_KEYS:
        return True
    if any(last_part.startswith(p) for p in WILDCARD_PREFIXES):
        return True
    if "eagle3" in key.lower():
        return True
    return False


# ============================================================
# 从日志中提取 non-default args 的原始文本
# ============================================================

def extract_raw_dict_text(log_path):
    """
    从日志中提取 non-default args 后的字典原始文本。
    支持：
    - 完整多行 dict（花括号匹配）
    - 单行内联 dict（可能被截断）
    """
    content = Path(log_path).read_text(encoding="utf-8", errors="replace")
    lines = content.split('\n')

    for line in lines:
        if 'non-default args:' not in line:
            continue
        idx = line.index('non-default args:')
        after = line[idx + len('non-default args:'):].strip()
        brace_idx = after.find('{')
        if brace_idx < 0:
            continue
        raw = after[brace_idx:]

        # 尝试完整花括号匹配（跨行）
        if raw.count('{') == raw.count('}'):
            return raw

        # 单行内：找到最后一个能让 depth=0 的位置，截取到那里
        depth = 0
        last_valid_end = 0
        for i, ch in enumerate(raw):
            if ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    last_valid_end = i + 1
        if last_valid_end > 0:
            # 截取到最后一个合法位置
            return raw[:last_valid_end]

        # 完全无法匹配 → 用整个 line 的 rest 作 raw text
        return raw

    return None


# ============================================================
# 方案A: 完整解析 (ast.literal_eval)
# ============================================================

def parse_args_string(s: str):
    if not s:
        return None
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        pass
    try:
        parsed = ast.literal_eval(s)
        if isinstance(parsed, dict):
            return parsed
    except (SyntaxError, ValueError):
        pass
    try:
        cleaned = s.strip()
        cleaned = re.sub(r',\s*}', '}', cleaned)
        cleaned = re.sub(r',\s*]', ']', cleaned)
        parsed = ast.literal_eval(cleaned)
        if isinstance(parsed, dict):
            return parsed
    except (SyntaxError, ValueError):
        pass
    return None


# ============================================================
# 方案B: 正则逐 key 提取（处理含 class 实例的截断 dict）
# ============================================================

def _match_python_value(text: str, pos: int):
    """
    从 text 的 pos 位置开始，尝试匹配一个 Python 值。
    返回 (value_str, end_pos) 或 None。
    """
    if pos >= len(text):
        return None

    ch = text[pos]
    rest = text[pos:]

    # None / True / False
    for kw in ('None', 'True', 'False'):
        if rest.startswith(kw) and (pos + len(kw) >= len(text) or text[pos + len(kw)] in ',} \t\n'):
            return (kw, pos + len(kw))

    # 数字
    num_m = re.match(r'-?\d+(?:\.\d+)?(?:e[+-]?\d+)?', rest)
    if num_m:
        return (num_m.group(), pos + num_m.end())

    # 字符串（单引号或双引号）
    if ch in ("'", '"'):
        quote = ch
        i = pos + 1
        while i < len(text):
            if text[i] == '\\':
                i += 2
                continue
            if text[i] == quote:
                return (text[pos:i + 1], i + 1)
            i += 1
        return None

    # 列表 [...]
    if ch == '[':
        depth = 0
        i = pos
        while i < len(text):
            if text[i] == '[':
                depth += 1
            elif text[i] == ']':
                depth -= 1
                if depth == 0:
                    return (text[pos:i + 1], i + 1)
            elif text[i] in ("'", '"'):
                # 跳过字符串内的内容
                quote = text[i]
                i += 1
                while i < len(text):
                    if text[i] == '\\':
                        i += 2
                        continue
                    if text[i] == quote:
                        break
                    i += 1
            i += 1
        return None

    # dict {...} 或 class 实例 xxx(...)
    if ch == '{':
        depth = 0
        i = pos
        while i < len(text):
            if text[i] == '{':
                depth += 1
            elif text[i] == '}':
                depth -= 1
                if depth == 0:
                    return (text[pos:i + 1], i + 1)
            elif text[i] in ("'", '"'):
                quote = text[i]
                i += 1
                while i < len(text):
                    if text[i] == '\\':
                        i += 2
                        continue
                    if text[i] == quote:
                        break
                    i += 1
            i += 1
        return None

    # ClassName(...) - Python object identity representation
    obj_m = re.match(r'[A-Za-z_][\w.]*\(', rest)
    if obj_m:
        # 找到匹配的 )
        paren_depth = 0
        i = pos
        while i < len(text):
            if text[i] == '(':
                paren_depth += 1
            elif text[i] == ')':
                paren_depth -= 1
                if paren_depth == 0:
                    return (text[pos:i + 1], i + 1)
            elif text[i] in ("'", '"'):
                quote = text[i]
                i += 1
                while i < len(text):
                    if text[i] == '\\':
                        i += 2
                        continue
                    if text[i] == quote:
                        break
                    i += 1
            i += 1
        # 未闭合 -> 截断，仍返回
        return (text[pos:], len(text))

    return None


def extract_interested_by_regex(raw_text: str) -> dict:
    """
    用正则从 raw dict 文本中提取所有关注的 key-value 对。
    不依赖完整解析，即使 dict 被截断也能提取已有内容。
    """
    if not raw_text:
        return {}

    results = {}

    # 逐 key 扫描：找到 'key': 然后匹配值
    # 先尝试精确匹配关注的 key
    all_interested = set(INTERESTED_KEYS)
    # 对于通配前缀，先匹配通配，再添加具体 key
    wildcard_re = re.compile(r"'(" + WILDCARD_PREFIX_PAT + r"\w+)'")

    # 合并所有感兴趣 key 的正则
    key_pat = re.compile(
        r"'([^']+)'\s*:\s*"
    )

    pos = 0
    while pos < len(raw_text):
        m = key_pat.search(raw_text, pos)
        if not m:
            break
        key = m.group(1)
        val_pos = m.end()

        # 是否关注此 key
        interested = False
        if key in all_interested:
            interested = True
        elif any(key.startswith(p) for p in WILDCARD_PREFIXES):
            interested = True
        elif "eagle3" in key.lower():
            interested = True

        if interested:
            result = _match_python_value(raw_text, val_pos)
            if result:
                val_str, next_pos = result
                try:
                    results[key] = ast.literal_eval(val_str)
                except (SyntaxError, ValueError):
                    results[key] = val_str
                pos = next_pos
                continue

        # 不关注，或者值匹配失败 -> 跳过这个 key 继续
        pos = m.end()

    return results


# ============================================================
# 展平工具
# ============================================================

def extract_flat_key_value(data: dict, parent_key: str = ""):
    """把嵌套 dict 展平为 key-value 对。"""
    items = []
    for k, v in data.items():
        full_key = f"{parent_key}.{k}" if parent_key else str(k)
        if isinstance(v, dict):
            items.append((full_key, str(v)))
            items.extend(extract_flat_key_value(v, full_key))
        elif isinstance(v, (list, tuple)):
            items.append((full_key, str(v)))
        else:
            items.append((full_key, str(v)))
    return items


# ============================================================
# 主流程
# ============================================================

def process_log(log_path: str):
    """主处理函数。返回生成的 Excel 路径，或 None。"""
    log_file = Path(log_path)
    if not log_file.exists():
        print(f"❌ 文件不存在: {log_path}")
        return None

    print(f"🔍 读取日志: {log_file}")

    raw_text = extract_raw_dict_text(log_path)
    if raw_text is None:
        print("⚠️  未找到 'non-default args:' 内容")
        return None

    print(f"📄 找到 non-default args, 文本长度: {len(raw_text)} chars")

    # 方案A: 完整解析
    merged = parse_args_string(raw_text)
    if merged:
        print(f"  ✅ 完整解析成功: {len(merged)} 个字段")
    else:
        # 方案B: 正则逐 key
        print("  ⚠️  完整解析失败，切换到正则逐 key 提取...")
        merged = extract_interested_by_regex(raw_text)
        if merged:
            print(f"  ✅ 正则提取到 {len(merged)} 个参数")
        else:
            print("❌  无法提取任何参数")
            return None

    # 按关注过滤 + 展平
    all_items = extract_flat_key_value(merged)
    interested = [(k, v) for k, v in all_items if is_interested_key(k)]

    if not interested:
        print("⚠️  未找到关注的参数")
        return None

    print(f"📊 关注参数: {len(interested)} 个")

    # ============================================================
    # 分类顺序
    # ============================================================
    category_order = {
        "模型路径": ["model"],
        "量化方式": ["quantization"],
        "并行策略": ["tensor_parallel_size", "enable_expert_parallel",
                      "data_parallel_size", "pipeline_parallel_size"],
        "配置参数": ["max_model_len", "max_num_batched_tokens", "max_num_seqs"],
    }

    def get_category(key: str) -> str:
        for cat, keys in category_order.items():
            if key in keys:
                return cat
        return "特性开关"

    # ============================================================
    # 写入 Excel
    # ============================================================
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Non-default Args"

    headers = ["分类", "参数名", "参数值"]
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for cat in ["模型路径", "量化方式", "并行策略", "配置参数", "特性开关"]:
        for key, val in interested:
            if cat == "特性开关":
                cat_match = (get_category(key) == "特性开关")
            else:
                cat_match = (key in category_order[cat])
            if cat_match:
                ws.cell(row=row, column=1, value=cat).border = thin_border
                c2 = ws.cell(row=row, column=2, value=key)
                c2.border = thin_border
                c2.font = Font(name='Consolas', size=10)
                ws.cell(row=row, column=3, value=str(val)).border = thin_border
                row += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 80
    ws.freeze_panes = 'A2'

    # 输出到桌面
    desktop = Path.home() / "Desktop"
    output_path = desktop / f"{log_file.stem}_non_default_args.xlsx"
    wb.save(str(output_path))
    print(f"\n✅ Excel 已生成: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 extract_log_params.py <日志文件路径>")
        sys.exit(1)
    process_log(sys.argv[1])
